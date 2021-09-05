#!/usr/bin/env python
# -*- coding: utf-8 -*-


import openpyxl
from util import get_path


# 读取excel测试数据

class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        test_data = []
        # 获取excel数据
        for item in range(2, sheet.max_row + 1):
            sub_data = {}
            sub_data['case_id'] = sheet.cell(item, 1).value
            sub_data['title'] = sheet.cell(item, 2).value
            sub_data['method'] = sheet.cell(item, 3).value
            sub_data['headers'] = sheet.cell(item, 4).value
            sub_data['url'] = sheet.cell(item, 5).value
            sub_data['data'] = sheet.cell(item, 6).value
            sub_data['excepted'] = sheet.cell(item, 7).value

            test_data.append(sub_data)


        return test_data

    def write_back(self, row, result, testResult):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, 8).value = result
        sheet.cell(row, 9).value = testResult
        wb.save(self.file_name)


if __name__ == '__main__':
    res = DoExcel(get_path.test_excel_path, 'login').get_data()
    print(res)
